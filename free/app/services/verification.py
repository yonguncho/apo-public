# -*- coding: utf-8 -*-
"""
verification.py
===============
판정 정확도 검증 하네스 (Phase 2).

APO의 판정이 실제 장비 상태와 맞는지 사람이 표본 대조할 수 있게 돕는다.
세 부분으로 나뉜다.

  1) stratified_sample()   등급별 층화 표본 추출
  2) build_worksheet()     대조용 체크리스트 (확인 항목 + FortiGate CLI 명령어)
  3) compute_precision()   대조 결과를 넣으면 등급별 정확도 리포트

**이 모듈은 장비에 접속하지 않는다.** 명령어를 적어 줄 뿐이고, 실행과 대조는
사람이 한다. 자동 대조는 오히려 검증의 의미를 없앤다 — APO가 스스로 채점하면
APO의 해석 오류를 그대로 정답으로 삼게 된다.

기존 판정 코드는 건드리지 않는다. 이 모듈은 판정 결과를 읽기만 한다.
"""
from __future__ import annotations

import csv
import io
import random
from collections import Counter, defaultdict
from typing import Any, Iterable

# 워크시트 컬럼. 앞쪽은 APO가 채우고, 뒤 3개는 사람이 채운다.
WORKSHEET_COLUMNS = [
    ("policy_id",     "정책 ID"),
    ("policy_type",   "종류"),
    ("name",          "정책 이름"),
    ("urgency",       "APO 등급"),
    ("risk_level",    "위험도"),
    ("reason",        "APO 판정 근거"),
    ("recommended",   "권고 조치"),
    ("check_items",   "확인할 것"),
    ("cli_commands",  "실행할 명령어"),
    ("actual",        "실제 확인 결과 (사람이 입력)"),
    ("verdict",       "판정 일치? 일치/불일치/보류 (사람이 입력)"),
    ("note",          "비고 (사람이 입력)"),
]

VERDICT_MATCH = "일치"
VERDICT_MISMATCH = "불일치"
VERDICT_HOLD = "보류"


# ---------------------------------------------------------------------------
# 1) 층화 표본 추출
# ---------------------------------------------------------------------------

def stratified_sample(
    policies: Iterable[dict],
    per_level_min: int = 5,
    per_level_max: int = 10,
    total_cap: int = 50,
    seed: int = 20260729,
) -> list[dict]:
    """등급별로 고르게 표본을 뽑는다.

    등급이 편중된 실데이터에서 무작위로 뽑으면 다수 등급만 검증하게 된다.
    등급마다 최소 per_level_min개를 확보해, 건수가 적은 등급의 정확도도
    측정할 수 있게 한다.

    같은 등급 안에서는 '판정 근거'가 서로 다른 정책을 우선 고른다. 같은 사유
    5건을 보는 것보다 5가지 사유를 보는 편이 규칙을 더 많이 검증한다.

    seed를 고정해 재실행 시 같은 표본이 나오게 한다(대조 작업 중 표본이
    바뀌면 진행 중인 검증이 무의미해진다).
    """
    rng = random.Random(seed)
    by_level: dict[Any, list[dict]] = defaultdict(list)
    for p in policies:
        by_level[p.get("urgency")].append(p)

    picked: list[dict] = []
    for level in sorted(by_level, key=lambda x: (x is None, x)):
        rows = list(by_level[level])
        rng.shuffle(rows)

        # 사유가 다른 것부터 채운다
        seen_reasons: set[str] = set()
        chosen: list[dict] = []
        for r in rows:
            key = _reason_key(r.get("reason"))
            if key not in seen_reasons:
                seen_reasons.add(key)
                chosen.append(r)
            if len(chosen) >= per_level_max:
                break
        # 사유 종류가 모자라면 남은 것으로 최소치를 채운다
        if len(chosen) < per_level_min:
            for r in rows:
                if r not in chosen:
                    chosen.append(r)
                if len(chosen) >= per_level_min:
                    break
        picked.extend(chosen[:per_level_max])

    if len(picked) > total_cap:
        # 등급별 최소치는 지키면서 상한에 맞춰 줄인다
        trimmed: list[dict] = []
        counts: Counter = Counter()
        for p in sorted(picked, key=lambda x: (counts[x.get("urgency")], str(x.get("urgency")))):
            if len(trimmed) >= total_cap:
                break
            trimmed.append(p)
            counts[p.get("urgency")] += 1
        picked = trimmed
    return picked


def _reason_key(reason: Any) -> str:
    """숫자를 뺀 사유 문자열. 'Hit 13 < 500'과 'Hit 7 < 500'을 같은 종류로 본다."""
    s = str(reason or "")
    return "".join(ch for ch in s if not ch.isdigit())[:60]


# ---------------------------------------------------------------------------
# 2) 검증 워크시트
# ---------------------------------------------------------------------------

def _policy_show_command(policy: dict) -> str:
    pid = str(policy.get("policy_id", "")).strip()
    ptype = (policy.get("policy_type") or policy.get("type") or "firewall").lower()
    table = "firewall proxy-policy" if "proxy" in ptype else "firewall policy"
    return f"show {table} {pid}"


def _runtime_command(policy: dict) -> str:
    """사용량 확인. APO의 런타임 임포트가 이 출력 형식을 그대로 읽는다."""
    pid = str(policy.get("policy_id", "")).strip()
    return f"diagnose firewall iprope show 00100004 {pid}"


def derive_checks(policy: dict) -> tuple[list[str], list[str]]:
    """판정 근거에 따라 '확인할 것'과 '실행할 명령어'를 만든다.

    무엇을 근거로 그 등급이 나왔는지에 따라 확인 지점이 달라진다. 전부 같은
    명령어를 주면 사람이 무엇을 봐야 할지 알 수 없다.
    """
    reason = str(policy.get("reason") or "")
    checks: list[str] = []
    cmds: list[str] = [_policy_show_command(policy)]

    def add(check: str, cmd: str | None = None):
        if check not in checks:
            checks.append(check)
        if cmd and cmd not in cmds:
            cmds.append(cmd)

    low = reason.lower()

    if "disabled" in low:
        add("정책이 실제로 비활성(status disable) 상태인가")
    if "schedule expired" in low or "sched_age" in low:
        add("스케줄 만료일이 실제로 지났는가. 이름만 날짜이고 always는 아닌가")
    if "hit=0" in low or "unused" in low:
        add("누적 접속 횟수가 정말 0인가. 장비 재시작·이중화 전환으로 카운터가 "
            "초기화된 이력은 없는가", _runtime_command(policy))
    if "last used" in low or "hit " in low or "low hit" in low:
        add("마지막 사용 시각이 실제로 그 시점인가", _runtime_command(policy))
    if "risky service" in low:
        add("문제로 지목된 평문 프로토콜이 실제로 이 정책에 포함되는가")
        add("해당 서비스만 빼도 업무에 지장이 없는가 (담당자 확인 필요)")
        cmds.append("show firewall service custom <서비스명>")
        cmds.append("show firewall service group <그룹명>")
    if "any" in low or "all" in low:
        add("출발지/목적지/서비스가 실제로 all 인가. 이름만 'all'인 객체는 아닌가")
        cmds.append("show firewall address all")
    if "object" in low or "admin policy" in low:
        add("판정에 쓰인 예외 객체가 실제로 이 정책에 포함되는가")
        cmds.append("show firewall addrgrp <그룹명>")
    if "valid its" in low:
        add("정책 이름의 티켓 번호가 실제 승인 건과 일치하는가 (ITSM 대조)")
    if "noticket" in low.replace(" ", "") or "no ticket" in low:
        add("이 정책이 정말 승인 없이 만들어졌는가. 티켓 번호가 이름에 없을 뿐 "
            "별도 승인 기록이 있는 경우가 흔하다 (ITSM 대조)")
    if "temp" in low:
        add("이름의 '임시' 표기가 실제로 임시 정책이라는 뜻인가. 이름만 그렇고 "
            "상시 운영 중인 경우가 있다 (담당자 확인 필요)")

    # 판정이 어떤 근거 위에 서 있는지를 검증자에게 알린다.
    # 사용량 데이터 없이 이름만으로 내린 판정은 근거가 약한데, 그 사실이
    # 등급 숫자에는 드러나지 않는다.
    if "csv not loaded" in low:
        add("[주의] 이 판정은 사용량 데이터 없이 정책 이름만으로 내려졌다. "
            "실제 사용 중인지 반드시 장비에서 확인할 것", _runtime_command(policy))
    if "reg date unknown" in low or "age=none" in low:
        add("[주의] 등록일을 알 수 없어 사용 기간 판정이 생략됐다. "
            "실제 생성 시점을 확인할 것")
    if "s-u" in low or "s-s" in low or "server-user" in low or "server-server" in low:
        add("출발지/목적지가 실제로 사용자 대역/서버 대역이 맞는가")
        cmds.append("show firewall address <객체명>")
    if "deny" in low:
        add("action이 실제로 deny 인가")
    if "icmp" in low:
        add("허용 서비스가 ICMP 계열뿐인가")

    if not checks:
        checks.append("APO 판정 근거가 실제 설정과 맞는가")

    return checks, cmds


def build_rows(samples: Iterable[dict]) -> list[dict]:
    rows = []
    for p in samples:
        checks, cmds = derive_checks(p)
        rows.append({
            "policy_id":   p.get("policy_id"),
            "policy_type": p.get("policy_type") or p.get("type") or "firewall",
            "name":        p.get("name"),
            "urgency":     p.get("urgency"),
            "risk_level":  p.get("risk_level"),
            "reason":      p.get("reason"),
            "recommended": p.get("recommended_action"),
            "check_items": "\n".join(f"{i}. {c}" for i, c in enumerate(checks, 1)),
            "cli_commands": "\n".join(cmds),
            "actual": "",
            "verdict": "",
            "note": "",
        })
    return rows


def to_csv(rows: list[dict]) -> bytes:
    """UTF-8 BOM CSV (Excel 한글 호환)."""
    from app.services.remediation_service import _csv_safe

    out = io.StringIO()
    w = csv.writer(out)
    w.writerow([label for _, label in WORKSHEET_COLUMNS])
    for r in rows:
        w.writerow([_csv_safe(r.get(key, "")) for key, _ in WORKSHEET_COLUMNS])
    return out.getvalue().encode("utf-8-sig")


def to_xlsx(rows: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    from app.services.workbook_exporter import SEVERITY_FILLS, write_text_cell

    wb = Workbook()
    ws = wb.active
    ws.title = "Verification"

    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(color="FFFFFF", bold=True)
    for col, (_, label) in enumerate(WORKSHEET_COLUMNS, 1):
        c = ws.cell(row=1, column=col, value=label)
        c.fill, c.font = header_fill, header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r_i, row in enumerate(rows, 2):
        fill = SEVERITY_FILLS.get(row.get("urgency"), SEVERITY_FILLS[0])
        for c_i, (key, _) in enumerate(WORKSHEET_COLUMNS, 1):
            val = row.get(key, "")
            cell = write_text_cell(ws, r_i, c_i, "" if val is None else str(val))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            # 사람이 채울 칸은 색을 칠하지 않아 눈에 띄게 둔다
            if key not in ("actual", "verdict", "note"):
                cell.fill = fill

    widths = {"policy_id": 10, "policy_type": 10, "name": 28, "urgency": 9,
              "risk_level": 10, "reason": 40, "recommended": 30,
              "check_items": 46, "cli_commands": 44, "actual": 30,
              "verdict": 16, "note": 24}
    for c_i, (key, _) in enumerate(WORKSHEET_COLUMNS, 1):
        ws.column_dimensions[ws.cell(row=1, column=c_i).column_letter].width = widths.get(key, 18)
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# 3) 정밀도 계산
# ---------------------------------------------------------------------------

def compute_precision(rows: Iterable[dict]) -> dict:
    """사람이 채운 워크시트를 받아 등급별 정확도를 낸다.

    '보류'와 미기입은 분모에서 뺀다. 확인하지 못한 건을 정답이나 오답으로
    치면 수치가 실제보다 좋아지거나 나빠진다.
    """
    per_level: dict[Any, Counter] = defaultdict(Counter)
    for r in rows:
        v = str(r.get("verdict") or "").strip()
        level = r.get("urgency")
        if v == VERDICT_MATCH:
            per_level[level]["match"] += 1
        elif v == VERDICT_MISMATCH:
            per_level[level]["mismatch"] += 1
        elif v == VERDICT_HOLD:
            per_level[level]["hold"] += 1
        else:
            per_level[level]["unfilled"] += 1

    levels = {}
    tot_m = tot_x = 0
    for level, c in per_level.items():
        judged = c["match"] + c["mismatch"]
        levels[level] = {
            "match": c["match"], "mismatch": c["mismatch"],
            "hold": c["hold"], "unfilled": c["unfilled"],
            "judged": judged,
            "precision": (c["match"] / judged) if judged else None,
        }
        tot_m += c["match"]
        tot_x += c["mismatch"]

    overall_judged = tot_m + tot_x
    return {
        "levels": dict(sorted(levels.items(), key=lambda kv: (kv[0] is None, kv[0]))),
        "overall": {
            "match": tot_m, "mismatch": tot_x, "judged": overall_judged,
            "precision": (tot_m / overall_judged) if overall_judged else None,
        },
        # 표본이 적으면 수치를 신뢰하기 어렵다는 점을 호출측이 알 수 있게 남긴다
        "weak_levels": [lv for lv, d in levels.items()
                        if d["judged"] and d["judged"] < 5],
    }


def format_precision_report(result: dict) -> str:
    lines = ["=" * 64, "  APO 판정 정확도 리포트", "=" * 64, ""]
    lines.append(f"{'등급':<6}{'일치':>6}{'불일치':>8}{'보류':>6}{'미기입':>8}{'정확도':>10}")
    lines.append("-" * 64)
    for lv, d in result["levels"].items():
        p = "—" if d["precision"] is None else f"{d['precision']*100:.1f}%"
        lines.append(f"{str(lv):<6}{d['match']:>6}{d['mismatch']:>8}{d['hold']:>6}{d['unfilled']:>8}{p:>10}")
    o = result["overall"]
    lines.append("-" * 64)
    p = "—" if o["precision"] is None else f"{o['precision']*100:.1f}%"
    lines.append(f"{'전체':<6}{o['match']:>6}{o['mismatch']:>8}{'':>6}{'':>8}{p:>10}")
    if o["judged"] == 0:
        lines += ["", "대조 결과가 아직 입력되지 않았습니다."]
    if result["weak_levels"]:
        lines += ["", f"주의: 대조 건수가 5건 미만인 등급 {result['weak_levels']} 은 "
                      "수치를 신뢰하기 어렵습니다."]
    return "\n".join(lines)
