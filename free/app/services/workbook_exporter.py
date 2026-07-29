from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HEADER_FILL = PatternFill('solid', fgColor='0F172A')
HEADER_FONT = Font(color='FFFFFF', bold=True)
THIN_GRAY = Side(style='thin', color='D9E2EC')
FILTER_FILL = PatternFill('solid', fgColor='F8FAFC')


def build_workbook(path: str | Path, sheets: dict[str, Any]) -> Path:
    path = Path(path)
    wb = Workbook()
    wb.remove(wb.active)

    ordered = [
        'firewall_policy',
        'firewall_proxy_policy',
        'firewall_address',
        'firewall_addrgrp',
        'firewall_proxy_address',
        'firewall_proxy_addrgrp',
        'firewall_service_custom',
        'firewall_service_group',
        'system_interface',
        'parse_warnings',
    ]

    for key in ordered:
        sheet_data = sheets.get(key)
        if not sheet_data:
            continue
        _add_sheet(wb, sheet_data.get('title') or key, sheet_data.get('headers') or [], sheet_data.get('rows') or [])

    if not wb.sheetnames:
        ws = wb.create_sheet('Export')
        ws['A1'] = 'No data'

    wb.save(path)
    return path


def _add_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    safe_title = title[:31] if title else 'Sheet'
    ws = wb.create_sheet(safe_title)
    ws.freeze_panes = 'A2'
    ws.sheet_view.showGridLines = False

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(bottom=THIN_GRAY)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = write_text_cell(ws, row_idx, col_idx, _stringify(value))
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    if headers and rows:
        end_col = get_column_letter(len(headers))
        end_row = len(rows) + 1
        ref = f'A1:{end_col}{end_row}'
        table = Table(displayName=_table_name_from_title(safe_title), ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name='TableStyleMedium2',
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    for i, header in enumerate(headers, start=1):
        max_len = max([len(str(header))] + [len(str(r[i - 1] if i - 1 < len(r) else '')) for r in rows[:500]])
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 12), 40)

    ws.row_dimensions[1].height = 24


def _table_name_from_title(title: str) -> str:
    base = ''.join(ch for ch in title.title() if ch.isalnum())
    if not base:
        base = 'Sheet'
    return f'Tbl{base[:20]}'


def _stringify(value: Any) -> Any:
    if isinstance(value, (list, tuple, set)):
        return " | ".join(str(v) for v in value)
    if isinstance(value, dict):
        return str(value)
    return value


# 스프레드시트가 수식으로 해석하는 선두 문자.
# 리포트에 들어가는 값(정책명·주석·주소 객체명 등)은 전부 분석 대상 config에서
# 온 것이라 신뢰할 수 없다. 예: 정책명이 "=cmd|'/C calc'!A1" 이면 openpyxl이
# 이를 실제 수식 셀로 저장하고, 분석가가 파일을 열 때 DDE가 실행된다.
FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")


def is_formula_like(value: Any) -> bool:
    """수식으로 해석될 수 있는 값인지.

    선두 문자만 보고 전부 막으면 오탐이 크다. 실제로 이 리포트에는
    last_used="-" 같은 '데이터 없음' 표시가 1500건 넘게 들어가는데, 그것까지
    중화하면 리포트가 지저분해지고 값이 왜곡된다. 단독 기호와 순수 숫자는
    스프레드시트가 수식으로 실행할 수 없으므로 그대로 둔다.
    """
    if not isinstance(value, str) or not value.startswith(FORMULA_TRIGGERS):
        return False
    if len(value) == 1:          # "-", "@" 같은 단독 기호
        return False
    try:
        float(value)             # "-5", "+3.2" 같은 순수 숫자
        return False
    except ValueError:
        return True


def write_text_cell(ws, row: int, column: int, value: Any):
    """셀에 값을 쓰되, 수식으로 해석될 값은 문자열로 강제한다.

    아포스트로피를 덧붙이는 흔한 방식 대신 data_type을 문자열로 고정한다.
    화면에 보이는 값이 원본 그대로 유지돼야 리포트가 왜곡되지 않기 때문이다.
    (저장·재로드 후에도 data_type='s'가 유지되는 것을 확인했다.)
    """
    cell = ws.cell(row=row, column=column, value=value)
    if is_formula_like(value):
        cell.data_type = "s"
    return cell


from io import BytesIO

SEVERITY_FILLS = {
    0: PatternFill("solid", fgColor="F0EEE7"),
    1: PatternFill("solid", fgColor="FFCCCC"),
    2: PatternFill("solid", fgColor="D3D1C7"),
    3: PatternFill("solid", fgColor="FFE0B2"),
    4: PatternFill("solid", fgColor="B5D4F4"),
    5: PatternFill("solid", fgColor="FFF9C4"),
    6: PatternFill("solid", fgColor="C0DD97"),
    7: PatternFill("solid", fgColor="9FE1CB"),
}

SEVERITY_COLS = [
    ("urgency",             "Severity"),
    ("risk_level",          "Risk Level"),
    # 등급(위험도)과 조치(무엇을 할 것인가)는 다른 축이다. 등급만으로 정렬하면
    # 성격이 다른 작업이 섞이므로 조치 유형을 별도 컬럼으로 낸다.
    ("action_label",        "Action"),
    ("recommended_action",  "Recommended Action"),
    ("reason",              "Reason"),
    ("traffic_type",        "Traffic Type"),
    ("tags",                "Tags"),
    ("policy_id",           "Policy ID"),
    ("name",                "Policy Name"),
    ("srcaddr_display",     "Source Address"),
    ("dstaddr_display",     "Destination Address"),
    ("service_display",     "Service"),
    ("action",              "Action"),
    ("status",              "Status"),
    ("schedule",            "Schedule"),
    ("hit_count",           "Hit Count"),
    ("last_used",           "Last Used"),
]


def build_severity_workbook(result: dict) -> bytes:
    """result: {"firewall": [...], "proxy": [...]} -> xlsx bytes"""
    wb = Workbook()
    wb.remove(wb.active)

    all_policies = result.get("firewall", []) + result.get("proxy", [])

    sheet_map = [
        ("Severity_All",    all_policies),
        ("Firewall_Policy", result.get("firewall", [])),
        ("Proxy_Policy",    result.get("proxy", [])),
    ]

    for sheet_name, policies in sheet_map:
        ws = wb.create_sheet(sheet_name)
        for col_idx, (field, label) in enumerate(SEVERITY_COLS, 1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row_idx, p in enumerate(policies, 2):
            sev = p.get("urgency", 0)
            fill = SEVERITY_FILLS.get(sev, SEVERITY_FILLS[0])
            for col_idx, (field, label) in enumerate(SEVERITY_COLS, 1):
                val = p.get(field, "")
                if isinstance(val, list):
                    val = ", ".join(str(v) for v in val)
                cell = write_text_cell(ws, row_idx, col_idx, str(val) if val else "")
                cell.fill = fill
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        ws.freeze_panes = "A2"

    _add_notes_sheet(wb, result)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _add_notes_sheet(wb: Workbook, result: dict) -> None:
    """리포트를 읽기 전에 알아야 할 것들을 첫 시트로 넣는다.

    판정 결과만 주고 한계를 말하지 않으면, 읽는 사람이 수치를 실제보다 확정적인
    것으로 받아들인다. 특히 Hit Count는 장비 재시작으로 초기화되므로 '낮음'이
    '미사용'을 뜻하지 않는다.
    """
    from app.services.severity_engine import KNOWN_LIMITATIONS

    ws = wb.create_sheet("Notes", 0)
    title_font = Font(bold=True, size=12)
    head_font = Font(bold=True)

    r = 1
    ws.cell(row=r, column=1, value="APO Analysis Report — Read This First").font = title_font
    r += 2

    ws.cell(row=r, column=1, value="Limitations of this report").font = head_font
    r += 1
    for item in KNOWN_LIMITATIONS:
        c = write_text_cell(ws, r, 1, f"• {item}")
        c.alignment = Alignment(vertical="top", wrap_text=True)
        r += 1
    r += 1

    inactive = result.get("inactive_rules") or []
    if inactive:
        ws.cell(row=r, column=1, value="Checks not applied in this analysis").font = head_font
        r += 1
        for item in inactive:
            c = write_text_cell(
                ws, r, 1,
                f"• {item.get('item')} — {item.get('reason')} {item.get('effect')}")
            c.alignment = Alignment(vertical="top", wrap_text=True)
            r += 1
        r += 1

    ws.cell(row=r, column=1, value="What the Action column means").font = head_font
    r += 1
    for label, desc in (
        ("Disable now",            "Disable this policy now."),
        ("Review candidate",       "Candidate for removal — confirm first. Already disabled or past its expiry date."),
        ("Disable & monitor",      "Disable first, watch for 30–90 days, then decide whether to remove."),
        ("Remove service only",    "Keep the policy; remove only the flagged service from it."),
        ("Needs review",           "Needs a closer look before deciding."),
        ("Register ticket",        "Keep the policy, but register it through your approval process."),
        ("No risk",                "Assessed and found not to be a risk (deny rules, ICMP-only, and similar)."),
        ("Not assessed (exempted)", "An exception rule stopped this policy from being assessed. This does NOT mean it is safe."),
        ("Cannot assess",          "Not enough information to judge."),
    ):
        write_text_cell(ws, r, 1, label)
        write_text_cell(ws, r, 2, desc)
        r += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 100
