from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HEADER_FILL = PatternFill('solid', fgColor='0F172A')
HEADER_FONT = Font(color='FFFFFF', bold=True)
THIN_GRAY = Side(style='thin', color='D9E2EC')
FILTER_FILL = PatternFill('solid', fgColor='F8FAFC')


def build_workbook(path: str | Path, sheets: dict[str, Any]) -> Path:
    path = Path(path)
    wb = Workbook()
    wb.remove(wb.active)

    ordered = [
        'firewall_policy',
        'firewall_proxy_policy',
        'firewall_address',
        'firewall_addrgrp',
        'firewall_proxy_address',
        'firewall_proxy_addrgrp',
        'firewall_service_custom',
        'firewall_service_group',
        'system_interface',
        'parse_warnings',
    ]

    for key in ordered:
        sheet_data = sheets.get(key)
        if not sheet_data:
            continue
        _add_sheet(wb, sheet_data.get('title') or key, sheet_data.get('headers') or [], sheet_data.get('rows') or [])

    if not wb.sheetnames:
        ws = wb.create_sheet('Export')
        ws['A1'] = 'No data'

    wb.save(path)
    return path


def _add_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    safe_title = title[:31] if title else 'Sheet'
    ws = wb.create_sheet(safe_title)
    ws.freeze_panes = 'A2'
    ws.sheet_view.showGridLines = False

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(bottom=THIN_GRAY)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_stringify(value))
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    if headers and rows:
        end_col = get_column_letter(len(headers))
        end_row = len(rows) + 1
        ref = f'A1:{end_col}{end_row}'
        table = Table(displayName=_table_name_from_title(safe_title), ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name='TableStyleMedium2',
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    for i, header in enumerate(headers, start=1):
        max_len = max([len(str(header))] + [len(str(r[i - 1] if i - 1 < len(r) else '')) for r in rows[:500]])
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 12), 40)

    ws.row_dimensions[1].height = 24


def _table_name_from_title(title: str) -> str:
    base = ''.join(ch for ch in title.title() if ch.isalnum())
    if not base:
        base = 'Sheet'
    return f'Tbl{base[:20]}'


def _stringify(value: Any) -> Any:
    if isinstance(value, (list, tuple, set)):
        return " | ".join(str(v) for v in value)
    if isinstance(value, dict):
        return str(value)
    return value


from io import BytesIO

SEVERITY_FILLS = {
    0: PatternFill("solid", fgColor="F1EFE8"),
    1: PatternFill("solid", fgColor="FFCCCC"),
    2: PatternFill("solid", fgColor="D3D1C7"),
    3: PatternFill("solid", fgColor="FFE0B2"),
    4: PatternFill("solid", fgColor="B5D4F4"),
    5: PatternFill("solid", fgColor="FFF9C4"),
    6: PatternFill("solid", fgColor="C0DD97"),
    7: PatternFill("solid", fgColor="9FE1CB"),
}

SEVERITY_COLS = [
    "urgency", "risk_level", "recommended_action", "reason",
    "traffic_type", "tags",
    "policy_id", "name", "ritm", "request_date", "requester",
    "srcaddr_display", "dstaddr_display", "service_display",
    "action", "status", "schedule", "hit_count", "last_used",
]


def build_severity_workbook(result: dict) -> bytes:
    """result: {"firewall": [...], "proxy": [...], "multicast": [...]} -> xlsx bytes"""
    wb = Workbook()
    wb.remove(wb.active)

    all_policies = (result.get("firewall", [])
                  + result.get("proxy", [])
                  + result.get("multicast", []))

    sheet_map = [
        ("Severity_All",     all_policies),
        ("Firewall_Policy",  result.get("firewall", [])),
        ("Proxy_Policy",     result.get("proxy", [])),
        ("Multicast_Policy", result.get("multicast", [])),
    ]

    for sheet_name, policies in sheet_map:
        ws = wb.create_sheet(sheet_name)
        for col_idx, col in enumerate(SEVERITY_COLS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row_idx, p in enumerate(policies, 2):
            sev = p.get("urgency", 0)
            fill = SEVERITY_FILLS.get(sev, SEVERITY_FILLS[0])
            for col_idx, col in enumerate(SEVERITY_COLS, 1):
                val = p.get(col, "")
                if isinstance(val, list):
                    val = ", ".join(str(v) for v in val)
                cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val else "")
                cell.fill = fill
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
